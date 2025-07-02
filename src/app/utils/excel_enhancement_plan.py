# ==========================================================
# RIEPILOGO CONTENUTO FILE:
# - Classe principale: AdvancedExcelExporter
# - Funzioni: enhance_existing_excel_system, test_advanced_excel_export
# ==========================================================

# ========================================
# EXCEL ENHANCEMENT - PIANO DETTAGLIATO
# Ricreazione funzionalità avanzate del file originale
# ========================================

"""
ANALISI DEL FILE ORIGINALE:
- ValutazioneCertificatesMemoria_v7.xlsx aveva:
  ✅ Grafici performance
  ✅ Risk metrics visualizzati  
  ✅ Analisi scenari multipli
  ✅ Dashboard professionale
  ✅ Formattazione avanzata

OBIETTIVO: Ricreare tutto questo in modo automatico
"""

# ========================================
# PARTE 1: ADVANCED EXCEL EXPORTER
# ========================================

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
import io
import numpy as np
from datetime import datetime, timedelta

class AdvancedExcelExporter:
    """Sistema avanzato per export Excel con grafici e dashboard"""
    
    def __init__(self, output_path="D:/Doc/File python/"):
        self.output_path = output_path
        self.workbook = None
        self.styles = self._create_styles()
        
        # Configurazione grafica
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
        
    def _create_styles(self):
        """Crea stili professionali per Excel"""
        return {
            'header': Font(bold=True, size=14, color="FFFFFF"),
            'subheader': Font(bold=True, size=12, color="2F5597"),
            'data': Font(size=10),
            'header_fill': PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid"),
            'alt_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin')),
            'center': Alignment(horizontal='center', vertical='center')
        }
    
    def create_comprehensive_report(self, certificate, analysis_results, scenarios=None):
        """Crea report Excel completo come quello originale"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CertificateAnalysis_Comprehensive_{timestamp}.xlsx"
        full_path = f"{self.output_path}{filename}"
        
        print(f"🚀 Creazione report Excel completo: {filename}")
        
        self.workbook = Workbook()
        
        # Rimuovi sheet default
        self.workbook.remove(self.workbook.active)
        
        # Crea tutti i sheet
        self._create_dashboard_sheet(certificate, analysis_results)
        self._create_risk_analysis_sheet(certificate, analysis_results)
        self._create_scenario_analysis_sheet(certificate, scenarios or self._generate_scenarios())
        self._create_performance_sheet(certificate, analysis_results)
        self._create_charts_sheet(certificate, analysis_results)
        self._create_data_sheet(certificate, analysis_results)
        
        # Salva file
        self.workbook.save(full_path)
        print(f"✅ Report salvato: {full_path}")
        return full_path
    
    def _create_dashboard_sheet(self, certificate, analysis_results):
        """Sheet 1: Dashboard Esecutivo"""
        
        ws = self.workbook.create_sheet("📊 Dashboard", 0)
        
        # Header principale
        ws.merge_cells('A1:H1')
        ws['A1'] = f"CERTIFICATE ANALYSIS DASHBOARD - {certificate.specs.name}"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws['A1'].alignment = self.styles['center']
        
        # Informazioni certificate
        self._add_certificate_info_table(ws, certificate, start_row=3)
        
        # Key metrics summary
        self._add_key_metrics_summary(ws, analysis_results, start_row=3, start_col=5)
        
        # Risk indicators
        self._add_risk_indicators(ws, analysis_results, start_row=12)
        
        # Formattazione colonne
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _create_risk_analysis_sheet(self, certificate, analysis_results):
        """Sheet 2: Analisi Rischio Dettagliata"""
        
        ws = self.workbook.create_sheet("⚠️ Risk Analysis")
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "RISK ANALYSIS - DETAILED METRICS"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws['A1'].alignment = self.styles['center']
        
        # Risk metrics table
        risk_data = self._prepare_risk_data(analysis_results)
        self._create_formatted_table(ws, risk_data, start_row=3, title="Risk Metrics")
        
        # VaR breakdown
        var_data = self._prepare_var_breakdown(analysis_results)
        self._create_formatted_table(ws, var_data, start_row=15, title="VaR Analysis")
        
        # Crea grafico VaR embedded
        self._add_var_chart(ws, analysis_results, anchor='H3')
    
    def _create_scenario_analysis_sheet(self, certificate, scenarios):
        """Sheet 3: Analisi Scenari"""
        
        ws = self.workbook.create_sheet("🔮 Scenarios")
        
        # Header
        ws.merge_cells('A1:G1')
        ws['A1'] = "SCENARIO ANALYSIS"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws['A1'].alignment = self.styles['center']
        
        # Scenarios table
        scenario_data = self._prepare_scenario_data(scenarios)
        self._create_formatted_table(ws, scenario_data, start_row=3, title="Scenario Results")
        
        # Stress test summary
        stress_data = self._prepare_stress_test_data(scenarios)
        self._create_formatted_table(ws, stress_data, start_row=15, title="Stress Test Summary")
    
    def _create_performance_sheet(self, certificate, analysis_results):
        """Sheet 4: Performance Analysis"""
        
        ws = self.workbook.create_sheet("📈 Performance")
        
        # Simula dati performance storici
        performance_data = self._simulate_performance_data(certificate)
        
        # Tabella performance
        self._create_performance_table(ws, performance_data)
        
        # Grafici performance
        self._add_performance_charts(ws, performance_data)
    
    def _create_charts_sheet(self, certificate, analysis_results):
        """Sheet 5: Grafici Dettagliati"""
        
        ws = self.workbook.create_sheet("📊 Charts")
        
        # Crea grafici matplotlib e inseriscili
        charts = [
            self._create_risk_distribution_chart(analysis_results),
            self._create_scenario_comparison_chart(analysis_results),
            self._create_payoff_profile_chart(certificate),
            self._create_sensitivity_analysis_chart(certificate)
        ]
        
        # Posiziona grafici nel sheet
        positions = ['A1', 'I1', 'A25', 'I25']
        for chart, position in zip(charts, positions):
            if chart:
                img = Image(chart)
                img.anchor = position
                ws.add_image(img)
    
    def _create_data_sheet(self, certificate, analysis_results):
        """Sheet 6: Dati Grezzi"""
        
        ws = self.workbook.create_sheet("📋 Raw Data")
        
        # Export tutti i dati per trasparenza
        raw_data = self._prepare_raw_data(certificate, analysis_results)
        
        for i, (table_name, data) in enumerate(raw_data.items()):
            start_row = i * 20 + 1
            self._create_formatted_table(ws, data, start_row, table_name)
    
    # ========================================
    # HELPER METHODS per dati e formattazione
    # ========================================
    
    def _add_certificate_info_table(self, ws, certificate, start_row):
        """Aggiunge tabella info certificato"""
        
        info_data = {
            'Field': ['Name', 'ISIN', 'Type', 'Issue Date', 'Maturity', 'Strike', 'Current Spot'],
            'Value': [
                certificate.specs.name,
                certificate.specs.isin,
                certificate.specs.certificate_type,
                certificate.specs.issue_date.strftime('%Y-%m-%d'),
                certificate.specs.maturity_date.strftime('%Y-%m-%d'),
                f"€{certificate.specs.strike:,.2f}",
                f"€{certificate.get_current_spot():,.2f}" if hasattr(certificate, 'get_current_spot') else 'N/A'
            ]
        }
        
        # Header
        ws[f'A{start_row}'] = "CERTIFICATE INFORMATION"
        ws[f'A{start_row}'].font = self.styles['subheader']
        
        # Dati
        for i, (field, value) in enumerate(zip(info_data['Field'], info_data['Value'])):
            row = start_row + i + 1
            ws[f'A{row}'] = field
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
    
    def _add_key_metrics_summary(self, ws, analysis_results, start_row, start_col):
        """Aggiunge summary metriche chiave"""
        
        col_letter = chr(ord('A') + start_col - 1)
        
        ws[f'{col_letter}{start_row}'] = "KEY METRICS"
        ws[f'{col_letter}{start_row}'].font = self.styles['subheader']
        
        # Fair value
        fair_value = analysis_results.get('fair_value', {})
        risk_metrics = analysis_results.get('risk_metrics', {})
        
        metrics = [
            ('Fair Value', f"€{fair_value.get('fair_value', 0):,.2f}"),
            ('Expected Return', f"{fair_value.get('expected_return', 0):.2%}"),
            ('VaR 95%', f"{risk_metrics.get('var_95', 0):.2%}"),
            ('Volatility', f"{risk_metrics.get('volatility', 0):.2%}"),
            ('Sharpe Ratio', f"{risk_metrics.get('sharpe_ratio', 0):.3f}")
        ]
        
        for i, (metric, value) in enumerate(metrics):
            row = start_row + i + 1
            ws[f'{col_letter}{row}'] = metric
            ws[f'{chr(ord(col_letter) + 1)}{row}'] = value
    
    def _create_formatted_table(self, ws, data, start_row, title):
        """Crea tabella formattata"""
        
        # Titolo
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = self.styles['subheader']
        
        # Header
        headers = list(data.keys())
        for i, header in enumerate(headers):
            cell = ws.cell(start_row + 1, i + 1, header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
        
        # Dati
        max_len = max(len(values) for values in data.values())
        for row in range(max_len):
            for col, values in enumerate(data.values()):
                if row < len(values):
                    cell = ws.cell(start_row + 2 + row, col + 1, values[row])
                    cell.font = self.styles['data']
                    if row % 2 == 0:
                        cell.fill = self.styles['alt_fill']
    
    def _prepare_risk_data(self, analysis_results):
        """Prepara dati risk per tabella"""
        
        risk_metrics = analysis_results.get('risk_metrics', {})
        
        return {
            'Metric': ['VaR 95%', 'VaR 99%', 'CVaR 95%', 'CVaR 99%', 'Volatility', 'Max Drawdown'],
            'Value': [
                f"{risk_metrics.get('var_95', 0):.2%}",
                f"{risk_metrics.get('var_99', 0):.2%}",
                f"{risk_metrics.get('cvar_95', 0):.2%}",
                f"{risk_metrics.get('cvar_99', 0):.2%}",
                f"{risk_metrics.get('volatility', 0):.2%}",
                f"{risk_metrics.get('maximum_loss', 0):.2%}"
            ],
            'Risk Level': ['High', 'Very High', 'High', 'Very High', 'Medium', 'High']
        }
    
    def _create_risk_distribution_chart(self, analysis_results):
        """Crea grafico distribuzione rischi"""
        
        fig, ax = plt.subplots(figsize=(8, 6))
        
        # Simula distribuzione returns
        np.random.seed(42)
        returns = np.random.normal(0.05, 0.15, 1000)
        
        ax.hist(returns, bins=50, alpha=0.7, color='skyblue', edgecolor='black')
        ax.axvline(np.percentile(returns, 5), color='red', linestyle='--', label='VaR 95%')
        ax.set_title('Return Distribution Analysis', fontsize=14, fontweight='bold')
        ax.set_xlabel('Returns')
        ax.set_ylabel('Frequency')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Salva come immagine
        img_path = f"{self.output_path}risk_distribution.png"
        plt.savefig(img_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return img_path
    
    def _generate_scenarios(self):
        """Genera scenari di default per analisi"""
        
        return {
            'base_case': {'return': 0.05, 'volatility': 0.20, 'probability': 0.50},
            'bull_case': {'return': 0.15, 'volatility': 0.25, 'probability': 0.25},
            'bear_case': {'return': -0.10, 'volatility': 0.35, 'probability': 0.25}
        }
    
    # Placeholder per altri metodi...
    def _prepare_var_breakdown(self, analysis_results):
        return {'Component': ['Market Risk', 'Credit Risk', 'Liquidity Risk'], 
                'VaR Contribution': ['€1,000', '€200', '€100']}
    
    def _add_var_chart(self, ws, analysis_results, anchor):
        pass  # Implementazione chart Excel nativo
    
    def _prepare_scenario_data(self, scenarios):
        return {'Scenario': list(scenarios.keys()), 
                'Return': [f"{s['return']:.2%}" for s in scenarios.values()]}
    
    def _prepare_stress_test_data(self, scenarios):
        return {'Test': ['Market Crash', 'Rate Shock'], 'Impact': ['-15%', '-8%']}
    
    def _simulate_performance_data(self, certificate):
        # Simula dati performance
        dates = pd.date_range(start='2023-01-01', end='2024-12-31', freq='D')
        values = 100 * np.cumprod(1 + np.random.normal(0.0002, 0.01, len(dates)))
        return pd.DataFrame({'Date': dates, 'Value': values})
    
    def _create_performance_table(self, ws, performance_data):
        pass  # Implementazione tabella performance
    
    def _add_performance_charts(self, ws, performance_data):
        pass  # Implementazione grafici performance
    
    def _create_scenario_comparison_chart(self, analysis_results):
        return None  # Placeholder
    
    def _create_payoff_profile_chart(self, certificate):
        return None  # Placeholder
    
    def _create_sensitivity_analysis_chart(self, certificate):
        return None  # Placeholder
    
    def _prepare_raw_data(self, certificate, analysis_results):
        return {'Market Data': {'Date': ['2024-01-01'], 'Price': [100]}}

# ========================================
# INTEGRATION con sistema esistente
# ========================================

def enhance_existing_excel_system():
    """Integra il nuovo sistema con quello esistente"""
    
    # Modifica da fare in unified_demo_system.py
    code_modification = """
    # SOSTITUIRE la classe ExcelIntegration con:
    
    class ExcelIntegration:
        def __init__(self):
            self.logger = logger
            self.base_path = "D:/Doc/File python/"
            self.advanced_exporter = AdvancedExcelExporter(self.base_path)
            
        def export_certificate_analysis(self, certificate, analysis_results, filename=None):
            # Usa il nuovo sistema avanzato
            if filename and 'comprehensive' in filename.lower():
                return self.advanced_exporter.create_comprehensive_report(
                    certificate, analysis_results
                )
            else:
                # Mantieni sistema semplice per compatibilità
                return self._export_simple_analysis(certificate, analysis_results, filename)
    """
    
    return code_modification

# ========================================
# TESTING
# ========================================

def test_advanced_excel_export():
    """Test del nuovo sistema Excel"""
    
    print("🧪 Test Advanced Excel Export")
    print("="*40)
    
    # Simula dati di test
    from types import SimpleNamespace
    
    # Mock certificate
    certificate = SimpleNamespace()
    certificate.specs = SimpleNamespace()
    certificate.specs.name = "Test Express Certificate"
    certificate.specs.isin = "TEST123456"
    certificate.specs.certificate_type = "express"
    certificate.specs.issue_date = datetime(2024, 1, 1)
    certificate.specs.maturity_date = datetime(2026, 1, 1)
    certificate.specs.strike = 1000.0
    certificate.get_current_spot = lambda: 1050.0
    
    # Mock analysis results
    analysis_results = {
        'fair_value': {
            'fair_value': 1025.50,
            'expected_return': 0.0855
        },
        'risk_metrics': {
            'var_95': -0.12,
            'var_99': -0.18,
            'cvar_95': -0.15,
            'cvar_99': -0.22,
            'volatility': 0.25,
            'maximum_loss': -0.35,
            'sharpe_ratio': 0.85
        }
    }
    
    # Test export
    try:
        exporter = AdvancedExcelExporter()
        result = exporter.create_comprehensive_report(certificate, analysis_results)
        
        if result:
            print(f"✅ Test riuscito: {result}")
            return True
        else:
            print("❌ Test fallito")
            return False
            
    except Exception as e:
        print(f"❌ Errore test: {e}")
        return False

if __name__ == "__main__":
    test_advanced_excel_export()